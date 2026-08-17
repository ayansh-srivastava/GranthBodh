from __future__ import annotations

from collections import Counter
from io import BytesIO

import fitz
from docx import Document
from docx.oxml.ns import qn
from docx.table import Table as DocxTable
from docx.text.paragraph import Paragraph as DocxParagraph
from openpyxl import load_workbook

from app.modules.document.chunker import TokenAwareChunker
from app.modules.document.data_model import Element, DocumentBlock

# PDF extraction
class PdfExtractor:
    # PDF extraction using PyMuPDF.
    # Produces layout aware elements instead of one huge string per page.

    MIN_CHARS_TO_KEEP_PAGE = 20

    @classmethod
    def extract(cls, file_bytes: bytes) -> list[Element]:
        doc = fitz.open(
            stream=file_bytes,
            filetype="pdf",
        )

        elements: list[Element] = []

        page_full_texts = {
            i: doc[i].get_text("text")
            for i in range(len(doc))
        }

        header_footer_lines = cls._detect_repeated_lines(
            page_full_texts
        )

        for page_index in range(len(doc)):
            page = doc[page_index]
            full_text = page_full_texts[page_index]

            if len(full_text.strip()) < cls.MIN_CHARS_TO_KEEP_PAGE:
                continue

            try:
                found_tables = page.find_tables()
            except Exception:
                found_tables = []

            table_bboxes: list[fitz.Rect] = []

            for table in found_tables:
                try:
                    grid = table.extract()
                except Exception:
                    continue

                if not grid:
                    continue

                if not any(
                    any(cell for cell in row)
                    for row in grid
                ):
                    continue

                table_bboxes.append(
                    fitz.Rect(table.bbox)
                )

                table_text = cls._table_to_markdown(grid)

                if table_text:
                    elements.append(
                        Element(
                            text=table_text,
                            element_type="table",
                            location={
                                "page": page_index + 1,
                            },
                        )
                    )

            for block in page.get_text("blocks"):
                x0, y0, x1, y1, text, *_rest = block

                text = text.strip()

                if not text:
                    continue

                block_rect = fitz.Rect(
                    x0,
                    y0,
                    x1,
                    y1,
                )

                # Don't duplicate table text as normal paragraph text.
                if any(
                    block_rect.intersects(table_bbox)
                    for table_bbox in table_bboxes
                ):
                    continue

                lines = [
                    line.strip()
                    for line in text.split("\n")
                    if line.strip()
                ]

                lines = [
                    line
                    for line in lines
                    if line not in header_footer_lines
                ]

                if not lines:
                    continue

                elements.append(
                    Element(
                        text=" ".join(lines),
                        element_type="paragraph",
                        location={
                            "page": page_index + 1,
                        },
                    )
                )

        doc.close()

        return elements

    @staticmethod
    def _table_to_markdown(
        grid: list[list[str | None]],
    ) -> str:

        rows = [
            [
                "" if cell is None else str(cell).strip()
                for cell in row
            ]
            for row in grid
        ]

        rows = [
            row
            for row in rows
            if any(row)
        ]

        if not rows:
            return ""

        header, *body = rows

        lines = [
            "| " + " | ".join(header) + " |",
            "| " + " | ".join(
                ["---"] * len(header)
            ) + " |",
        ]

        for row in body:
            lines.append(
                "| " + " | ".join(row) + " |"
            )

        return "\n".join(lines)

    @staticmethod
    def _detect_repeated_lines(
        page_texts: dict[int, str],
        min_page_fraction: float = 0.6,
    ) -> set[str]:

        if len(page_texts) < 3:
            return set()

        line_page_counts: Counter[str] = Counter()

        for text in page_texts.values():

            seen_this_page = {
                line.strip()
                for line in text.split("\n")
                if line.strip()
                and len(line.strip()) <= 80
            }

            line_page_counts.update(
                seen_this_page
            )

        threshold = max(
            3,
            int(
                len(page_texts)
                * min_page_fraction
            ),
        )

        return {
            line
            for line, count
            in line_page_counts.items()
            if count >= threshold
        }


# DOCX extraction
class DocxExtractor:

    @classmethod
    def extract(
        cls,
        file_bytes: bytes,
    ) -> list[Element]:

        document = Document(
            BytesIO(file_bytes)
        )

        elements: list[Element] = []

        # heading level -> heading text
        heading_stack: dict[int, str] = {}

        for block in cls._iter_block_items(document):

            if isinstance(
                block,
                DocxParagraph,
            ):

                text = block.text.strip()

                if not text:
                    continue

                style_name = (
                    block.style.name
                    if block.style
                    else "Normal"
                )

                if style_name.startswith("Heading"):

                    try:
                        level = int(
                            style_name.split()[-1]
                        )
                    except ValueError:
                        level = 1

                    heading_stack = {
                        key: value
                        for key, value
                        in heading_stack.items()
                        if key < level
                    }

                    section_path_before = [
                        heading_stack[key]
                        for key in sorted(
                            heading_stack
                        )
                    ]

                    heading_stack[level] = text

                    elements.append(
                        Element(
                            text=text,
                            element_type="heading",
                            level=level,
                            section_path=section_path_before,
                        )
                    )

                    continue

                section_path = [
                    heading_stack[key]
                    for key in sorted(
                        heading_stack
                    )
                ]

                elements.append(
                    Element(
                        text=text,
                        element_type="paragraph",
                        section_path=section_path,
                        location={
                            "style": style_name,
                        },
                    )
                )

            # Table
            elif isinstance(
                block,
                DocxTable,
            ):

                section_path = [
                    heading_stack[key]
                    for key in sorted(
                        heading_stack
                    )
                ]

                markdown = (
                    cls._table_to_markdown(block)
                )

                if markdown:
                    elements.append(
                        Element(
                            text=markdown,
                            element_type="table",
                            section_path=section_path,
                        )
                    )

        return elements

    @staticmethod
    def _iter_block_items(document: Document):

        parent_elm = document.element.body

        for child in parent_elm.iterchildren():

            if child.tag == qn("w:p"):
                yield DocxParagraph(
                    child,
                    document,
                )

            elif child.tag == qn("w:tbl"):
                yield DocxTable(
                    child,
                    document,
                )

    @staticmethod
    def _table_to_markdown(
        table: DocxTable,
    ) -> str:

        rows = [
            [
                cell.text.strip()
                for cell in row.cells
            ]
            for row in table.rows
        ]

        rows = [
            row
            for row in rows
            if any(row)
        ]

        if not rows:
            return ""

        header, *body = rows

        lines = [
            "| " + " | ".join(header) + " |",
            "| " + " | ".join(
                ["---"] * len(header)
            ) + " |",
        ]

        for row in body:
            lines.append(
                "| " + " | ".join(row) + " |"
            )

        return "\n".join(lines)


# XLSX extraction
class XlsxExtractor:

    HEADER_SCAN_ROWS = 5

    @classmethod
    def extract(
        cls,
        file_bytes: bytes,
    ) -> list[Element]:

        workbook = load_workbook(
            BytesIO(file_bytes),
            read_only=True,
            data_only=True,
        )

        elements: list[Element] = []

        for sheet_index, worksheet in enumerate(
            workbook.worksheets,
            start=1,
        ):

            all_rows = list(
                worksheet.iter_rows(
                    values_only=True
                )
            )

            if not all_rows:
                continue

            header_row_index, headers = (
                cls._find_header_row(all_rows)
            )

            if headers is None:
                continue

            elements.append(
                Element(
                    text=(
                        f"Sheet '{worksheet.title}' "
                        f"— columns: "
                        f"{', '.join(headers)}"
                    ),
                    element_type="heading",
                    level=1,
                    location={
                        "sheet_name": worksheet.title,
                        "sheet_index": sheet_index,
                    },
                )
            )

            # Each row becomes an element.
            for row_offset, row in enumerate(
                all_rows[
                    header_row_index + 1:
                ],
                start=header_row_index + 2,
            ):

                if all(
                    value is None
                    for value in row
                ):
                    continue

                pairs = []

                for col_idx, value in enumerate(
                    row
                ):

                    if col_idx >= len(headers):
                        break

                    value_str = (
                        ""
                        if value is None
                        else str(value).strip()
                    )

                    pairs.append(
                        f"{headers[col_idx]}: "
                        f"{value_str}"
                    )

                elements.append(
                    Element(
                        text=(
                            f"Row {row_offset} — "
                            + ", ".join(pairs)
                        ),
                        element_type="table",
                        location={
                            "sheet_name": worksheet.title,
                            "sheet_index": sheet_index,
                            "row_start": row_offset,
                            "row_end": row_offset,
                        },
                    )
                )

        return elements

    @classmethod
    def _find_header_row(
        cls,
        all_rows: list[tuple],
    ) -> tuple[
        int,
        list[str] | None,
    ]:

        scan_window = (
            all_rows[: cls.HEADER_SCAN_ROWS]
            or [[]]
        )

        max_cols = max(
            (
                len(row)
                for row in scan_window
            ),
            default=0,
        )

        if max_cols == 0:
            return 0, None

        for index, row in enumerate(
            scan_window
        ):

            non_empty = sum(
                1
                for value in row
                if value not in (
                    None,
                    "",
                )
            )

            if non_empty < max(
                2,
                max_cols // 2,
            ):
                continue

            has_data_after = any(
                sum(
                    1
                    for value in candidate
                    if value not in (
                        None,
                        "",
                    )
                ) >= 1
                for candidate in all_rows[
                    index + 1:
                    index + 3
                ]
            )

            if has_data_after:

                headers = [
                    (
                        str(value).strip()
                        if value is not None
                        else f"Column {idx + 1}"
                    )
                    for idx, value
                    in enumerate(row)
                ]

                return index, headers

        row = all_rows[0]

        headers = [
            (
                str(value).strip()
                if value is not None
                else f"Column {idx + 1}"
            )
            for idx, value
            in enumerate(row)
        ]

        return 0, headers


# Public DocumentParser
class DocumentParser:

    @staticmethod
    def parse_pdf(
        file_bytes: bytes,
        doc_id: str,
        filename: str,
        chunker: TokenAwareChunker | None = None,
    ) -> list[DocumentBlock]:

        elements = PdfExtractor.extract(
            file_bytes
        )

        if chunker is None:
            raise ValueError(
                "A TokenAwareChunker with a "
                "GeminiTokenCounter is required."
            )

        return chunker.chunk(
            elements,
            {
                "doc_id": doc_id,
                "filename": filename,
                "source_type": "pdf",
            },
        )

    @staticmethod
    def parse_docx(
        file_bytes: bytes,
        doc_id: str,
        filename: str,
        chunker: TokenAwareChunker | None = None,
    ) -> list[DocumentBlock]:

        elements = DocxExtractor.extract(
            file_bytes
        )

        if chunker is None:
            raise ValueError(
                "A TokenAwareChunker with a "
                "GeminiTokenCounter is required."
            )

        return chunker.chunk(
            elements,
            {
                "doc_id": doc_id,
                "filename": filename,
                "source_type": "docx",
            },
        )

    @staticmethod
    def parse_xlsx(
        file_bytes: bytes,
        doc_id: str,
        filename: str,
        chunker: TokenAwareChunker | None = None,
    ) -> list[DocumentBlock]:

        elements = XlsxExtractor.extract(
            file_bytes
        )

        if chunker is None:
            raise ValueError(
                "A TokenAwareChunker with a "
                "GeminiTokenCounter is required."
            )

        return chunker.chunk(
            elements,
            {
                "doc_id": doc_id,
                "filename": filename,
                "source_type": "xlsx",
            },
        )

    @classmethod
    def parse(
        cls,
        filename: str,
        file_bytes: bytes,
        doc_id: str | None = None,
        chunker: TokenAwareChunker | None = None,
    ) -> list[DocumentBlock]:

        extension = (
            filename
            .lower()
            .rsplit(".", 1)[-1]
        )

        doc_id = doc_id or filename

        if extension == "pdf":
            return cls.parse_pdf(
                file_bytes,
                doc_id,
                filename,
                chunker,
            )

        if extension == "docx":
            return cls.parse_docx(
                file_bytes,
                doc_id,
                filename,
                chunker,
            )

        if extension == "xlsx":
            return cls.parse_xlsx(
                file_bytes,
                doc_id,
                filename,
                chunker,
            )

        raise ValueError(
            f"Unsupported file type: .{extension}"
        )