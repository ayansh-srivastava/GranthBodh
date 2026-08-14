from io import BytesIO

from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader


from dataclasses import dataclass, field
from typing import Any


@dataclass
class DocumentBlock:
    text: str
    metadata: dict[str, Any] = field(default_factory=dict)

class DocumentParser:

    @staticmethod
    def parse_pdf(file_bytes: bytes) -> list[DocumentBlock]:
        reader = PdfReader(BytesIO(file_bytes))

        blocks = []

        for page_number, page in enumerate(reader.pages, start=1):
            text = page.extract_text()

            if text and text.strip():
                blocks.append(
                    DocumentBlock(
                        text=text.strip(),
                        metadata={
                            "page": page_number,
                        },
                    )
                )

        return blocks

    @staticmethod
    def parse_docx(file_bytes: bytes) -> list[DocumentBlock]:
        document = Document(BytesIO(file_bytes))

        blocks: list[DocumentBlock] = []

        heading_stack: dict[int, str] = {}

        for paragraph in document.paragraphs:
            text = paragraph.text.strip()

            if not text:
                continue

            style_name = paragraph.style.name if paragraph.style else "Normal"

            if style_name.startswith("Heading"):
                try:
                    level = int(style_name.split()[-1])
                except ValueError:
                    level = 1

                heading_stack = {
                    key: value
                    for key, value in heading_stack.items()
                    if key < level
                }

                heading_stack[level] = text

                continue

            section_path = [
                heading_stack[level]
                for level in sorted(heading_stack)
            ]

            blocks.append(
                DocumentBlock(
                    text=text,
                    metadata={
                        "type": "paragraph",
                        "section_path": section_path,
                        "section": " > ".join(section_path),
                        "style": style_name,
                    },
                )
            )

        return blocks

    @staticmethod
    def parse_xlsx(file_bytes: bytes) -> list[DocumentBlock]:
        workbook = load_workbook(
            BytesIO(file_bytes),
            read_only=True,
            data_only=True,
        )

        blocks = []

        MAX_ROWS_PER_BLOCK = 30

        for index, worksheet in enumerate(workbook.worksheets, start=1):
            rows_iterator = worksheet.iter_rows(values_only=True)

            # First row is treated as the header or title row
            header_row = next(rows_iterator, None)

            if not header_row:
                continue

            headers = [
                str(value).strip() if value is not None else f"Column {index + 1}"
                for index, value in enumerate(header_row)
            ]

            current_rows: list[str] = []
            chunk_start_row = 2

            for row_number, row in enumerate(rows_iterator, start=2):
                # skip completely empty rows
                if all(value is None for value in row):
                    continue

                row_lines = []

                for column_index, value in enumerate(row):
                    if column_index >= len(headers):
                        break

                    header = headers[column_index]

                    value = "" if value is None else str(value).strip()

                    row_lines.append(f"{header}: {value}")

                row_text = "\n".join(row_lines)

                current_rows.append(f"Row {row_number}:\n{row_text}")

                if len(current_rows) >= MAX_ROWS_PER_BLOCK:
                    blocks.append(
                        DocumentBlock(
                            text=(
                                f"Sheet: {worksheet.title}\n\n"
                                f"Columns: {' | '.join(headers)}\n\n"
                                f"{chr(10).join(current_rows)}"
                            ),
                            metadata={
                                "sheet_index": index,
                                "sheet_name": worksheet.title,
                                "row_start": chunk_start_row,
                                "row_end": row_number,
                            },
                        )
                    )

                    current_rows = []
                    chunk_start_row = row_number + 1

            if current_rows:
                blocks.append(
                    DocumentBlock(
                        text=(
                            f"Sheet: {worksheet.title}\n\n"
                            f"Columns: {' | '.join(headers)}\n\n"
                            f"{chr(10).join(current_rows)}"
                        ),
                        metadata={
                            "sheet_index": index,
                            "sheet_name": worksheet.title,
                            "row_start": chunk_start_row,
                            "row_end": chunk_start_row + len(current_rows) - 1,
                        },
                    )
                )

        return blocks

    @classmethod
    def parse( cls, filename: str, file_bytes: bytes) -> list[DocumentBlock]:

        extension = filename.lower().split(".")[-1]

        if extension == "pdf":
            return cls.parse_pdf(file_bytes)

        if extension == "docx":
            return cls.parse_docx(file_bytes)

        if extension == "xlsx":
            return cls.parse_xlsx(file_bytes)

        raise ValueError(f"Unsupported file type: .{extension}")
